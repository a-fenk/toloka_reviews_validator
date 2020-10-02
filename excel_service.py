from enum import Enum
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill


class Colors(Enum):
    WHITE = 'FFFFFF'
    GREY = 'A0A0A0'
    YELLOW = 'FFFF00'
    TURQUOISE = '40E0D0'
    GREEN = '00FF7F'
    RED = 'FF0000'
    BLUE = '6666FF'


class CellData:
    def __init__(self, color: Colors, row: int, column: str, data: str = None):
        if data:
            self.data = data
        self.color = color
        self.row = row
        self.column = column


def create_workbook():
    workbook = Workbook(iso_dates=True)
    workbook.remove(workbook["Sheet"])
    return workbook


def save_workbook(workbook):
    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    return virtual_workbook


def read_data_from_workbook(workbook: Workbook):
    sheets_names = workbook.get_sheet_names()
    sheets_data = []
    for sheet_name in sheets_names:
        sheet_data = {'name': sheet_name, 'data': []}
        sheet = workbook.get_sheet_by_name(sheet_name)
        row = 2
        while sheet['D' + str(row)].value:
            sheet_data['data'].append({
                'full_desc': sheet['D' + str(row)].value,
                'short_descr': sheet['E' + str(row)].value,
                'result_gold': sheet['I' + str(row)].value,
                'row': row
            })
            row += 1
        sheets_data.append(sheet_data)
    return sheets_data


def write_data_to_sheet(workbook: Workbook, sheet_name: str, data: list):
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet['U1'] = 'Dublicate short_desc'
    sheet['V1'] = 'Unique words result_gold'
    sheet['W1'] = 'Unique words full_desc'
    sheet['X1'] = 'Dublicate word result_gold'
    sheet['Y1'] = 'Dublicate word full_desc'
    sheet['Z1'] = 'Dublicate short_desc - Unique words result_gold'
    sheet['AA1'] = 'Dublicate word result_gold - Dublicate word full_desc'

    sheet['U1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['V1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['W1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['X1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['Y1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['Z1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)
    sheet['AA1'].fill = PatternFill("solid", fgColor=Colors.YELLOW.value)

    for elem in data:
        if hasattr(elem, 'data'):
            sheet[elem.column + str(elem.row)] = elem.data
        sheet[elem.column + str(elem.row)].fill = PatternFill("solid", fgColor=elem.color.value)
    return workbook
